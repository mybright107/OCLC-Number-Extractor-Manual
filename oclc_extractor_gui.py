import re
import sys
import traceback
import xml.etree.ElementTree as ET
from pathlib import Path
from tkinter import Tk, filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


APP_NAME = "OCLC Number Extractor"
COLUMNS = [
    "OCLC_Number",
    "Full_035_Field",
    "Record_Number",
    "Resource_Type",
]


def local_name(element):
    """Return an XML element name without its namespace."""
    return element.tag.split("}")[-1]


def find_available_output_path(input_path):
    """Create a new filename without overwriting an existing workbook."""
    first_path = input_path.with_name(
        f"{input_path.stem}_oclc_numbers.xlsx"
    )

    if not first_path.exists():
        return first_path

    counter = 2

    while True:
        candidate = input_path.with_name(
            f"{input_path.stem}_oclc_numbers_{counter}.xlsx"
        )

        if not candidate.exists():
            return candidate

        counter += 1


def extract_records(input_path):
    tree = ET.parse(input_path)
    root = tree.getroot()

    results = []
    record_number = 0

    for record in root.iter():
        if local_name(record) != "record":
            continue

        record_number += 1
        oclc_number = None
        full_oclc_identifier = None
        is_online = False

        for field in record.iter():
            if local_name(field) != "datafield":
                continue

            tag = field.attrib.get("tag", "")

            if tag == "035" and oclc_number is None:
                field_text = "".join(field.itertext())
                match = re.search(r"\(OCoLC\)(\d+)", field_text)

                if match:
                    oclc_number = match.group(1)
                    full_oclc_identifier = match.group(0)

            if tag == "300":
                field_text = " ".join(field.itertext())

                if re.search(
                    r"online resource",
                    field_text,
                    re.IGNORECASE,
                ):
                    is_online = True

        if oclc_number:
            results.append(
                {
                    "OCLC_Number": oclc_number,
                    "Full_035_Field": full_oclc_identifier,
                    "Record_Number": record_number,
                    "Resource_Type": (
                        "Online" if is_online else "Physical"
                    ),
                }
            )

    return results, record_number


def format_worksheet(worksheet):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 18,
        "B": 24,
        "C": 16,
        "D": 18,
    }

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    # Keep OCLC numbers as text so Excel does not alter them.
    for cell in worksheet["A"][1:]:
        cell.number_format = "@"


def add_worksheet(workbook, title, records):
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(COLUMNS)

    for record in records:
        worksheet.append(
            [record[column] for column in COLUMNS]
        )

    format_worksheet(worksheet)


def create_workbook(records, output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    online_records = [
        record
        for record in records
        if record["Resource_Type"] == "Online"
    ]

    physical_records = [
        record
        for record in records
        if record["Resource_Type"] == "Physical"
    ]

    add_worksheet(workbook, "All records", records)
    add_worksheet(workbook, "Online", online_records)
    add_worksheet(workbook, "Physical", physical_records)

    workbook.save(output_path)

    return len(online_records), len(physical_records)


def main():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    input_filename = filedialog.askopenfilename(
        title="Select the XML file",
        filetypes=[
            ("XML files", "*.xml"),
            ("All files", "*.*"),
        ],
    )

    if not input_filename:
        return

    input_path = Path(input_filename)

    try:
        records, total_records = extract_records(input_path)

        if not records:
            messagebox.showwarning(
                APP_NAME,
                "No OCLC numbers in the form "
                "'(OCoLC)123456' were found.",
            )
            return

        output_path = find_available_output_path(input_path)

        online_count, physical_count = create_workbook(
            records,
            output_path,
        )

        messagebox.showinfo(
            APP_NAME,
            "Processing completed successfully.\n\n"
            f"XML records examined: {total_records}\n"
            f"OCLC records found: {len(records)}\n"
            f"Online resources: {online_count}\n"
            f"Physical resources: {physical_count}\n\n"
            f"Excel file saved as:\n{output_path}",
        )

    except ET.ParseError as error:
        messagebox.showerror(
            APP_NAME,
            "The selected file could not be parsed as XML.\n\n"
            f"Details: {error}",
        )

    except PermissionError:
        messagebox.showerror(
            APP_NAME,
            "The output workbook could not be saved.\n\n"
            "Close the workbook if it is already open, then try again.",
        )

    except Exception as error:
        error_log = input_path.with_name(
            f"{input_path.stem}_error_log.txt"
        )

        try:
            error_log.write_text(
                traceback.format_exc(),
                encoding="utf-8",
            )
        except Exception:
            pass

        messagebox.showerror(
            APP_NAME,
            "The file could not be processed.\n\n"
            f"Error: {error}\n\n"
            f"An error log was written to:\n{error_log}",
        )


if __name__ == "__main__":
    main()